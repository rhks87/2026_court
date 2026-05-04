"""
화성 테니스 빈 코트 캘린더 v0.6
변경사항:
  - 시간 단축버튼: 전체/오전/오후/저녁(18~) + 세부 2시간 버튼 연동
  - 코트 필터 버튼 배경색 꽉 채우기
  - 같은 그룹 = 완전 동일한 색
"""
import requests, json, time, csv, io, re
from datetime import datetime, timezone, timedelta

API_URL  = "https://yeyak.hscity.go.kr/stadium/stadiumReserveUseList.do"
RESV_URL = "https://yeyak.hscity.go.kr/stadiumDetail.do?stadiumIdx="
HEADERS  = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://yeyak.hscity.go.kr/stadiumDetail.do",
}

def fetch(idx, year, month):
    try:
        r = requests.post(API_URL,
            data={"stadiumIdx": idx, "searchYear": str(year), "searchMonth": str(month)},
            headers=HEADERS, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [X] {idx} {year}-{month}: {e}")
        return None

def extract_empty(resp):
    if not resp:
        return []
    return [{"date": s.get("sorDate"),
             "begin": s.get("stadiumBeginHm"),
             "end":   s.get("stadiumEndHm")}
            for s in resp.get("useCntList", [])
            if s.get("applyStatusCd") is None]

JUKMI_SHEET = "1Hg_MniS8eEc-SADH-FKqZQK-AWaogLXD"
JUKMI_RESV  = f"https://docs.google.com/spreadsheets/d/{JUKMI_SHEET}/htmlview"
JUKMI_XLSX  = f"https://docs.google.com/spreadsheets/d/{JUKMI_SHEET}/export?format=xlsx"

def fetch_jukmi(year, month):
    """
    구글 시트 XLSX 파싱 → 죽미 빈자리 추출

    핵심 발견사항 (디버깅으로 확인):
    1. "06-08"~"12-14" 시간 라벨 → Excel이 날짜로 자동변환 (2025-06-08...)
       → datetime 타입 체크로 월=시작시간, 일=종료시간 복원
    2. 데이터행 col B가 헤더행과 merge → 헤더로 오인식
       → RAW 셀만 체크, 헤더 발견 즉시 다음행을 데이터행으로 처리
    3. 회색(FF999999) = 대관불가, 흰색(FFFFFFFF) + 빈값 = 빈자리
    """
    import calendar, io
    from datetime import datetime as _dt
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    # xlsx 다운로드
    try:
        res = requests.get(JUKMI_XLSX, timeout=30,
                           headers={"User-Agent": "Mozilla/5.0"})
        if res.status_code != 200:
            return []
        wb = load_workbook(io.BytesIO(res.content), data_only=True)
    except Exception as e:
        print(f"[X] 죽미 xlsx 로드 실패: {e}")
        return []

    # 시트 이름으로 찾기 (탭명: "26년 5월")
    sheet_name = f"{year - 2000}년 {month}월"
    if sheet_name not in wb.sheetnames:
        return []  # 탭 아직 없음 (20일 이전)
    ws = wb[sheet_name]

    # 머지 셀 마스터 맵
    merge_master = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_master[(r, c)] = (rng.min_row, rng.min_col)

    def cell_info(r, c):
        """머지 처리 포함 셀 정보 반환"""
        master = merge_master.get((r, c), (r, c))
        cell = ws.cell(row=master[0], column=master[1])
        val = str(cell.value).strip() if cell.value is not None else ""
        return val, _is_grey(cell)

    def _is_grey(cell):
        """회색(대관불가) 판별 — FF999999 기준"""
        if isinstance(cell, MergedCell):
            return False
        try:
            fill = cell.fill
            if not fill or fill.fill_type in (None, "none"):
                return False
            fg = fill.fgColor
            if fg.type == "rgb":
                rgb = fg.rgb
                if len(rgb) == 8 and rgb not in ("00000000", "FFFFFFFF"):
                    r = int(rgb[2:4], 16)
                    g = int(rgb[4:6], 16)
                    b = int(rgb[6:8], 16)
                    diff = max(abs(r-g), abs(g-b), abs(r-b))
                    bright = (r+g+b) / 3
                    return diff < 45 and 30 < bright < 245
        except Exception:
            pass
        return False

    max_day = calendar.monthrange(year, month)[1]
    TIME_PAT = re.compile(r"^(\d{2})-(\d{2})$")
    DAY_COLS = [3, 5, 7, 9, 11, 13, 15]  # openpyxl 1-indexed

    def extract_day(cell):
        """셀에서 일(day) 숫자 추출"""
        if isinstance(cell, MergedCell):
            return None
        v = cell.value
        if isinstance(v, _dt):
            return v.day             # 날짜객체: .day = 일수
        if isinstance(v, (int, float)):
            return int(v)            # 1.0 → 1
        if isinstance(v, str):
            m = re.match(r"^(\d{1,2})", v)
            if m:
                return int(m.group(1))   # "5(공휴일)" → 5
        return None

    def detect_time(raw_cell):
        """
        col B의 RAW 셀에서 시간 범위 추출
        - "06-08" 등은 Excel이 날짜로 변환: datetime(2025, 6, 8) → month=6, day=8
        - "18-20" 등은 변환 불가(월>12)라서 문자열로 저장
        """
        if isinstance(raw_cell, MergedCell):
            return None   # merge된 셀 = 데이터행 → 헤더 아님
        v = raw_cell.value
        if isinstance(v, _dt):
            sh, eh = v.month, v.day   # month=시작시간, day=종료시간
            if 0 <= sh <= 23 and 0 <= eh <= 23:
                return (f"{sh:02d}:00", f"{eh:02d}:00")
        if isinstance(v, str):
            mt = TIME_PAT.match(v.strip())
            if mt:
                return (f"{mt.group(1)}:00", f"{mt.group(2)}:00")
        return None

    col_day = {}
    slot_dict = {}  # {(date, begin): end} — 같은 날짜는 마지막 주 블록이 덮어씀
    ri = 1  # 1-indexed

    while ri <= ws.max_row:

        # ── 날짜행 감지 ──────────────────────────────────────
        tmp = {}
        for ci in DAY_COLS:
            raw = ws.cell(row=ri, column=ci)
            d = extract_day(raw)
            if d is not None and 1 <= d <= max_day:
                tmp[ci] = d

        if len(tmp) >= 2:
            col_day = tmp
            ri += 1
            continue

        # ── 헤더행 감지 (RAW 셀 직접 체크) ──────────────────
        # merge된 셀이면 None → 데이터행으로 판단, 헤더 아님
        raw_b = ws.cell(row=ri, column=2)
        tr = detect_time(raw_b)

        if tr and col_day:
            begin, end = tr
            data_ri = ri + 1  # 헤더 바로 다음 행 = 데이터행

            for ci, day in col_day.items():
                try:
                    ds = f"{year}-{month:02d}-{day:02d}"
                    v1_text, v1_grey = cell_info(data_ri, ci)
                    v2_text, v2_grey = cell_info(data_ri, ci + 1)

                    # 둘 다 회색이면 대관불가
                    if v1_grey and v2_grey:
                        continue

                    # 회색 아니고 비어있는 코트가 하나라도 있으면 빈자리
                    v1_ok = not v1_grey and v1_text == ""
                    v2_ok = not v2_grey and v2_text == ""

                    if v1_ok or v2_ok:
                        slot_dict[(ds, begin)] = end  # 최신 주 블록으로 덮어쓰기
                    else:
                        # 해당 날짜가 예약/불가로 확정되면 이전 오판을 제거
                        slot_dict.pop((ds, begin), None)
                except IndexError:
                    pass

            ri += 2  # 헤더 + 데이터 동시 skip
            continue

        ri += 1

    # dict → list 변환 (올바른 주 블록 데이터만 남김)
    slots = [{"date": k[0], "begin": k[1], "end": v}
             for k, v in slot_dict.items()]
    return sorted(slots, key=lambda s: (s["date"], s["begin"]))


def main():
    with open("stadiums.json", encoding="utf-8") as f:
        stadiums = json.load(f)

    KST = timezone(timedelta(hours=9))
    now = datetime.now(KST)
    y, m = now.year, now.month
    months = [(y, m)]
    nm, ny = m+1, y
    if nm > 12: nm, ny = 1, y+1
    months.append((ny, nm))

    print(f"\n[조회] {months[0][0]}-{months[0][1]:02d} + {months[1][0]}-{months[1][1]:02d}")
    print("=" * 60)

    result = []
    for s in stadiums:
        idx = s["idx"]
        print(f"  [{idx:>4s}] {s['name']:<14s}", end=" ")
        slots = []
        for yr, mo in months:
            d = fetch(idx, yr, mo)
            time.sleep(0.3)
            slots.extend(extract_empty(d))
        result.append({
            "idx": idx, "name": s["name"], "group": s.get("group",""),
            "url": RESV_URL + idx, "empty_slots": slots
        })
        print(f"빈자리 {len(slots):>3d}개")

    # 죽미 실내 테니스장 (구글 시트)
    print(f"  [죽미] 죽미 실내 테니스장   ", end=" ")
    jukmi_slots = []
    for yr, mo in months:
        jukmi_slots.extend(fetch_jukmi(yr, mo))
    result.append({
        "idx": "jukmi", "name": "죽미 실내", "group": "죽미",
        "url": JUKMI_RESV, "empty_slots": jukmi_slots
    })
    print(f"빈자리 {len(jukmi_slots):>3d}개")

    ts = now.strftime("%Y-%m-%d %H:%M")
    html = (HTML
            .replace("__DATA__",  json.dumps(result, ensure_ascii=False))
            .replace("__TIME__",  ts)
            .replace("__YEAR__",  str(months[0][0]))
            .replace("__MONTH__", str(months[0][1])))

    with open("tennis_court.html", "w", encoding="utf-8") as f:
        f.write(html)

    total = sum(len(c["empty_slots"]) for c in result)
    print("=" * 60)
    print(f"[OK] tennis_court.html 생성 / 총 빈자리 {total}개")
    print("→ 더블클릭해서 브라우저에서 열어보세요 🎾")


HTML = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>화성 빈 코트 v0.6</title>
<style>
:root{
  --bg:#f0f2f5;--card:#fff;--text:#1a1a2e;--muted:#6b7280;
  --border:#e2e8f0;--accent:#3b82f6;--hover:#f1f5f9;
  --sun:#ef4444;--sat:#3b82f6;--today-ring:#f59e0b;
  --shadow:0 1px 4px rgba(0,0,0,.08);
}
[data-theme=dark]{
  --bg:#0f172a;--card:#1e293b;--text:#e2e8f0;--muted:#64748b;
  --border:#334155;--accent:#60a5fa;--hover:#273449;
  --sun:#f87171;--sat:#60a5fa;--today-ring:#fbbf24;
  --shadow:0 1px 4px rgba(0,0,0,.3);
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Noto Sans KR",-apple-system,sans-serif;
  background:var(--bg);color:var(--text);
  min-height:100vh;padding:16px;
  transition:background .25s,color .25s}

.hdr{display:flex;justify-content:space-between;align-items:flex-start;
  flex-wrap:wrap;gap:10px;margin-bottom:14px}
.hdr h1{font-size:20px;font-weight:800;letter-spacing:-.5px}
.hdr h1 em{font-size:11px;font-weight:400;color:var(--muted);
  font-style:normal;margin-left:6px}
.hdr p{font-size:11px;color:var(--muted);margin-top:3px}
.hdr-r{display:flex;gap:6px}
.update-time{font-size:11px;font-weight:400;color:var(--muted);margin-left:8px}

/* 공통 버튼 */
.btn{padding:6px 12px;border:1px solid var(--border);background:var(--card);
  color:var(--text);border-radius:8px;cursor:pointer;font-size:12px;
  font-family:inherit;font-weight:600;transition:all .15s;line-height:1.4}
.btn:hover{border-color:var(--accent);color:var(--accent)}
.btn.on{background:var(--accent);color:#fff;border-color:var(--accent)}
.btn.icon{padding:6px 10px;font-size:15px}

/* 필터 */
.filters{background:var(--card);border:1px solid var(--border);
  border-radius:12px;padding:12px 14px;margin-bottom:12px;
  display:flex;flex-direction:column;gap:10px;box-shadow:var(--shadow)}
.fg{display:flex;align-items:center;gap:5px;flex-wrap:wrap}
.fg-lbl{font-size:10px;font-weight:700;color:var(--muted);
  text-transform:uppercase;letter-spacing:.8px;min-width:42px;flex-shrink:0}
.fg-div{width:1px;height:20px;background:var(--border);margin:0 4px}

/* 단축 버튼 — 조금 더 굵게 */
.btn-short{padding:7px 14px;font-size:13px}

/* 세부 시간 버튼 — 작게 */
.btn-detail{padding:5px 9px;font-size:11px;border-radius:6px}

/* ★ 코트 버튼 — 배경색 꽉 채우기 */
.btn-court{
  padding:7px 14px;font-size:13px;font-weight:700;
  border:none;border-radius:8px;cursor:pointer;
  color:#fff;font-family:inherit;
  transition:filter .15s,transform .1s;
  opacity:1;
}
.btn-court:hover{filter:brightness(1.1);transform:translateY(-1px)}
.btn-court.off{opacity:.35;filter:grayscale(.4)}

/* 월 네비 */
.mnav{display:flex;justify-content:center;align-items:center;
  gap:16px;margin-bottom:10px}
.mnav-title{font-size:19px;font-weight:800;min-width:150px;text-align:center}
.mbtn{width:34px;height:34px;border-radius:50%;
  border:1px solid var(--border);background:var(--card);color:var(--text);
  cursor:pointer;font-size:15px;display:flex;align-items:center;
  justify-content:center;transition:all .15s}
.mbtn:hover{background:var(--hover);border-color:var(--accent)}
.mbtn:disabled{opacity:.3;cursor:default}

/* 캘린더 */
.cal-wrap{background:var(--card);border:1px solid var(--border);
  border-radius:12px;overflow:hidden;box-shadow:var(--shadow)}
table.cal{width:100%;border-collapse:collapse;table-layout:fixed}
table.cal th{padding:10px 4px;font-size:12px;font-weight:700;
  color:var(--muted);background:var(--hover);
  border-bottom:1px solid var(--border)}
th.h-sun{color:var(--sun)}
th.h-sat{color:var(--sat)}
table.cal td{
  border:1px solid var(--border);
  vertical-align:top;
  height:128px;          /* 6개 슬롯 기준 고정 */
  width:14.28%;
  padding:4px;
  background:var(--card);
  overflow:hidden;       /* 넘쳐도 셀 크기 유지 */
}
td.empty{background:var(--hover);opacity:.45}
td.past{opacity:.4}
td.today{box-shadow:inset 0 0 0 2px var(--today-ring)}

.dnum{font-size:11px;font-weight:700;margin-bottom:2px;  /* 날짜 숫자 작게 */
  padding:1px 3px;display:inline-block;border-radius:4px}
.dnum.sun{color:var(--sun)}
.dnum.sat{color:var(--sat)}
td.today .dnum{background:var(--today-ring);color:#fff;border-radius:50%;
  width:19px;height:19px;font-size:10px;
  display:flex;align-items:center;justify-content:center;padding:0}

.holi{font-size:10px;color:var(--sun);font-weight:700;
  margin-bottom:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
td.holiday-bg{background:rgba(239,68,68,.07)!important}
[data-theme=dark] td.holiday-bg{background:rgba(248,113,113,.1)!important}

.slots{display:grid;grid-template-columns:1fr 1fr;gap:3px;overflow:hidden}
.slots.exp{grid-template-columns:1fr 1fr}

.slot{
  display:block;padding:4px 6px;border-radius:5px;
  cursor:pointer;text-decoration:none;
  color:#fff;font-size:11px;font-weight:700;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;
  transition:filter .12s,transform .1s;line-height:1.3;
}
.slot:hover{filter:brightness(1.12);transform:translateY(-1px)}
.sn-s{display:none}

.more-btn{
  grid-column:1/-1;font-size:11px;color:var(--accent);font-weight:700;
  padding:3px 5px;cursor:pointer;text-align:center;
  border:1px dashed var(--border);border-radius:5px;
  background:none;font-family:inherit;transition:background .12s;margin-top:1px;
}
.more-btn:hover{background:var(--hover)}

.legend{display:flex;flex-wrap:wrap;gap:12px;margin-top:12px;
  justify-content:center;font-size:12px;color:var(--muted)}
.leg-item{display:flex;align-items:center;gap:6px}
.leg-dot{width:14px;height:14px;border-radius:3px}

.summary{font-size:11px;color:var(--muted);text-align:center;margin-top:8px}

.tip{position:fixed;pointer-events:none;z-index:9999;
  background:#1e293b;color:#fff;font-size:12px;font-weight:500;
  padding:5px 10px;border-radius:7px;white-space:nowrap;
  box-shadow:0 4px 12px rgba(0,0,0,.25);opacity:0;transition:opacity .1s}

@media(max-width:700px){
  table.cal td{height:120px;padding:4px 3px}
  .slot{font-size:10px;padding:3px 4px}
  .dnum{font-size:12px}
  .slots{grid-template-columns:1fr}
  .slot-hint{display:none}
  .sn-f{display:none}
  .sn-s{display:inline}
}
</style>
</head>
<body data-theme="light">
<div class="tip" id="tip"></div>

<div class="hdr">
  <div>
    <h1>🎾 동탄·죽미 코트 예약현황 <em>v0.7</em> <span class="update-time">__TIME__ 기준</span></h1>
  </div>
  <div class="hdr-r">
    <button class="btn icon" onclick="toggleTheme()">🌙</button>
  </div>
</div>

<div class="filters">
  <!-- 시간 필터 -->
  <div class="fg">
    <span class="fg-lbl">⏰ 시간</span>
    <!-- 단축 버튼 -->
    <button class="btn btn-short f-short" data-s="all"     onclick="setShort('all')">🔍 전체</button>
    <button class="btn btn-short f-short" data-s="morning" onclick="setShort('morning')">🌅 오전 (~12)</button>
    <button class="btn btn-short f-short" data-s="afternoon" onclick="setShort('afternoon')">☀️ 오후 (12~18)</button>
    <button class="btn btn-short on f-short" data-s="evening" onclick="setShort('evening')">🌙 저녁 (18~) ⭐</button>
    <!-- 구분선 -->
    <div class="fg-div"></div>
    <!-- 세부 버튼 -->
    <button class="btn btn-detail f-t" data-h="6"  onclick="togT(6)">06-08</button>
    <button class="btn btn-detail f-t" data-h="8"  onclick="togT(8)">08-10</button>
    <button class="btn btn-detail f-t" data-h="10" onclick="togT(10)">10-12</button>
    <button class="btn btn-detail f-t" data-h="12" onclick="togT(12)">12-14</button>
    <button class="btn btn-detail f-t" data-h="14" onclick="togT(14)">14-16</button>
    <button class="btn btn-detail f-t" data-h="16" onclick="togT(16)">16-18</button>
    <button class="btn btn-detail on f-t" data-h="18" onclick="togT(18)">18-20</button>
    <button class="btn btn-detail on f-t" data-h="20" onclick="togT(20)">20-22</button>
  </div>
  <!-- 코트 필터 -->
  <div class="fg" id="gf">
    <span class="fg-lbl">📍 코트</span>
  </div>
</div>

<div class="mnav">
  <button class="mbtn" id="pb" onclick="goM(-1)">◀</button>
  <div class="mnav-title" id="mt"></div>
  <button class="mbtn" id="nb" onclick="goM(1)">▶</button>
</div>

<div class="cal-wrap"><div id="cal"></div></div>


<script>
const COURTS = __DATA__;
const T0 = new Date(); T0.setHours(0,0,0,0);

/* ★ 그룹 단일 색 (lightness 고정 50%) */
const GH = {금반저류지:215, 왕배산:145, 여울공원:340, 돌모루:275, 죽미:25};
function groupColor(g){ return `hsl(${GH[g]??0},65%,50%)`; }
function slotColor(c){ return groupColor(c.group||c.name); }

function shortNm(c){
  const n=(c.name.match(/(\d+)번/)||[])[1]||'';
  const m={금반저류지:'금반',왕배산:'왕배산',여울공원:'여울',돌모루:'돌모루',죽미:'죽미'};
  return (m[c.group]||c.group)+n;
}
function mobileNm(c){
  const n=(c.name.match(/(\d+)번/)||[])[1]||'';
  const m={금반저류지:'금반',왕배산:'왕배',여울공원:'여울',돌모루:'돌모루',죽미:'죽미'};
  return (m[c.group]||c.group)+n;
}

const HOLI={
  /* 2026년 확정 공휴일 */
  '2026-01-01':'신정',
  '2026-02-16':'설 연휴','2026-02-17':'설날','2026-02-18':'설 연휴',
  '2026-03-01':'삼일절(일)','2026-03-02':'삼일절 대체',
  '2026-05-05':'어린이날',
  '2026-05-24':'부처님오신날(일)','2026-05-25':'부처님오신날 대체',
  '2026-06-03':'지방선거',
  '2026-06-06':'현충일(토)',
  '2026-08-15':'광복절(토)','2026-08-17':'광복절 대체',
  '2026-09-24':'추석 연휴','2026-09-25':'추석','2026-09-26':'추석 연휴(토)',
  '2026-10-03':'개천절(토)','2026-10-05':'개천절 대체',
  '2026-10-09':'한글날',
  '2026-12-25':'성탄절',
};

const MONTHS=[];
{const y=parseInt("__YEAR__"),m=parseInt("__MONTH__");
 MONTHS.push({y,m});
 let ny=y,nm=m+1;if(nm>12){nm=1;ny=y+1;}
 MONTHS.push({y:ny,m:nm});}
let cur=0;

/* 단축 그룹 정의 */
const SHORT_MAP = {
  all:       [6,8,10,12,14,16,18,20],
  morning:   [6,8,10],
  afternoon: [12,14,16],
  evening:   [18,20],
};

/* 필터 상태 — 기본: 저녁 ON */
let fHours = new Set([18,20]);
const allGroups=[...new Set(COURTS.map(c=>c.group||c.name))];
let fGroups=new Set(allGroups);
const expanded=new Set();

/* ★ 코트 버튼 — 배경색 꽉 채우기 */
const gf=document.getElementById('gf');
allGroups.forEach(g=>{
  const b=document.createElement('button');
  b.className='btn-court'; b.dataset.v=g; b.textContent=g;
  b.style.background=groupColor(g);
  b.onclick=()=>togG(g);
  gf.appendChild(b);
});
/* 같은 줄 오른쪽 끝에 힌트 텍스트 */
const hint=document.createElement('span');
hint.className='slot-hint';
hint.style.cssText='margin-left:auto;font-size:11px;color:var(--muted);white-space:nowrap';
hint.innerHTML='✅ 빈 슬롯 클릭하면 예약페이지로 이동 &nbsp;·&nbsp; 📋 +N개 → 펼치기';
gf.appendChild(hint);

/* 단축 버튼: 해당 시간 셋 켜기/끄기 */
function setShort(s){
  const hours = SHORT_MAP[s];
  const allOn = hours.every(h=>fHours.has(h));
  // 모두 켜진 상태 → 끄기 / 아니면 → 켜기 (all도 동일하게 토글)
  if(allOn){ hours.forEach(h=>fHours.delete(h)); }
  else      { hours.forEach(h=>fHours.add(h)); }
  syncUI(); expanded.clear(); render();
}

/* 세부 버튼 개별 토글 */
function togT(h){
  fHours.has(h)?fHours.delete(h):fHours.add(h);
  syncUI(); expanded.clear(); render();
}

/* UI 동기화 (세부 버튼 + 단축 버튼 active 상태) */
function syncUI(){
  // 세부 버튼
  document.querySelectorAll('.f-t').forEach(b=>
    b.classList.toggle('on', fHours.has(parseInt(b.dataset.h))));
  // 단축 버튼 — 해당 그룹 시간이 모두 켜진 경우 active
  document.querySelectorAll('.f-short').forEach(b=>{
    const s=b.dataset.s;
    if(s==='all'){
      b.classList.toggle('on', SHORT_MAP.all.every(h=>fHours.has(h)));
    } else {
      b.classList.toggle('on', SHORT_MAP[s].every(h=>fHours.has(h)));
    }
  });
}

function togG(v){
  if(fGroups.has(v)) fGroups.delete(v);
  else fGroups.add(v);
  // 버튼 off 클래스
  document.querySelectorAll('.btn-court').forEach(b=>
    b.classList.toggle('off', !fGroups.has(b.dataset.v)));
  expanded.clear(); render();
}
function goM(d){
  const n=cur+d; if(n<0||n>=MONTHS.length)return;
  cur=n; expanded.clear(); render();
}

function ok(slot,court){
  if(!fGroups.has(court.group||court.name)) return false;
  if(!fHours.has(parseInt(slot.begin))) return false;
  if(new Date(slot.date)<T0) return false;
  return true;
}
function slotsOn(ds){
  const out=[];
  COURTS.forEach(c=>c.empty_slots.forEach(s=>{
    if(s.date===ds&&ok(s,c)) out.push({...s,court:c});
  }));
  return out.sort((a,b)=>a.begin.localeCompare(b.begin)||
                          a.court.idx.localeCompare(b.court.idx));
}

function toggleExp(ds){
  expanded.has(ds)?expanded.delete(ds):expanded.add(ds);
  render();
}

const MAX=6; /* 2열×3행 */

function buildSlots(slots,ds){
  const isExp=expanded.has(ds);
  const vis=isExp?slots:slots.slice(0,MAX);
  const rest=slots.length-MAX;
  let h=`<div class="slots${isExp?' exp':''}">`;
  vis.forEach(s=>{
    const col=slotColor(s.court);
    const sn=shortNm(s.court);
    const tip2=`${s.court.name}  ${s.begin}~${s.end}`;
    h+=`<a class="slot" href="${s.court.url}" target="_blank"
      style="background:${col}"
      onmouseenter="showTip(event,'${tip2.replace(/'/g,"\\'")}')"
      onmouseleave="hideTip()"
    ><span class='t'>${s.begin}</span> <span class='sn-f'>${sn}</span><span class='sn-s'>${mobileNm(s.court)}</span></a>`;
  });
  if(!isExp&&rest>0){
    h+=`<button class="more-btn" onclick="toggleExp('${ds}')">+${rest}개 더 보기 🔽</button>`;
  } else if(isExp&&slots.length>MAX){
    h+=`<button class="more-btn" onclick="toggleExp('${ds}')">접기 🔼</button>`;
  }
  h+='</div>';
  return h;
}

function render(){
  const {y,m}=MONTHS[cur];
  document.getElementById('mt').textContent=`📅 ${y}년 ${m}월`;
  document.getElementById('pb').disabled=(cur===0);
  document.getElementById('nb').disabled=(cur===MONTHS.length-1);

  const fd=new Date(y,m-1,1),ld=new Date(y,m,0);
  const sd=fd.getDay(),td=ld.getDate();

  let html=`<table class="cal"><thead><tr>
    <th class="h-sun">일</th><th>월</th><th>화</th><th>수</th>
    <th>목</th><th>금</th><th class="h-sat">토</th>
  </tr></thead><tbody>`;

  let day=1,row=0,shown=0;
  while(row<6){                  /* 항상 정확히 6행 */
    html+='<tr>';
    for(let c=0;c<7;c++){
      const isBlank = (row===0&&c<sd) || day>td;
      if(isBlank){
        html+='<td class="empty"></td>';
      } else {
        const ds=`${y}-${String(m).padStart(2,'0')}-${String(day).padStart(2,'0')}`;
        const dt=new Date(y,m-1,day);
        const isPast=dt<T0,isToday=dt.getTime()===T0.getTime();
        const slots=slotsOn(ds);
        shown+=slots.length;
        const dw=dt.getDay();
        const dc=dw===0?'sun':dw===6?'sat':'';
        let cls=''; if(isPast)cls='past'; if(isToday)cls='today';
        const holi=HOLI[ds]||'';
        html+=`<td class="${cls}${holi?' holiday-bg':''}">`;
        html+=`<div class="dnum ${dc}">${day}</div>`;
        if(holi) html+=`<div class="holi">${holi}</div>`;
        html+=buildSlots(slots,ds);
        html+=`</td>`;
        day++;
      }
    }
    html+='</tr>'; row++;
  }
  html+='</tbody></table>';
  document.getElementById('cal').innerHTML=html;


}

const tip=document.getElementById('tip');
function showTip(e,txt){tip.textContent=txt;tip.style.opacity='1';moveTip(e);}
document.addEventListener('mousemove',moveTip);
function moveTip(e){tip.style.left=(e.clientX+12)+'px';tip.style.top=(e.clientY-30)+'px';}
function hideTip(){tip.style.opacity='0';}
function toggleTheme(){
  document.body.dataset.theme=document.body.dataset.theme==='dark'?'light':'dark';}
function saveJson(){
  const blob=new Blob([JSON.stringify(COURTS,null,2)],{type:'application/json'});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download=`tennis_${new Date().toISOString().slice(0,10)}.json`;
  a.click();
}

// 초기 UI 동기화
syncUI();
render();
</script>
</body>
</html>"""

if __name__ == "__main__":
    main()
